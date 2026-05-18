"""Compare weld lengths between correct answer and auto-generated results."""
import openpyxl
from collections import defaultdict

CORRECT_FILE = '焊缝统计R3_auto(1).xlsx'
AUTO_FILE    = '焊缝统计_auto.xlsx'

def load_correct(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[wb.sheetnames[0]]   # first sheet = 焊缝统计 (correct answer)
    rows = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        if r[0] is None:
            continue
        # columns: 序号, 位置, hf, 长度, 备注, 零件1, 零件2, 构件
        hf_val = r[2]   # keep None for CJP
        rows.append({
            'pos':  r[1],
            'hf':   hf_val,
            'len':  round(float(r[3] or 0), 1),
            'ann':  r[4] or '',
            'p1':   str(r[5] or ''),
            'p2':   str(r[6] or ''),
            'comp': str(r[7] or ''),
        })
    return rows

def load_auto(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb['焊缝统计']
    rows = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        if r[1] is None:
            continue
        hf_val = r[2]   # keep None for CJP
        rows.append({
            'pos':  r[1],
            'hf':   hf_val,
            'len':  round(float(r[3] or 0), 1),
            'ann':  r[4] or '',
            'p1':   str(r[5] or ''),
            'p2':   str(r[6] or ''),
            'comp': str(r[7] or ''),
        })
    return rows

manual = load_correct(CORRECT_FILE)
script = load_auto(AUTO_FILE)

# Only compare components for which we have DXF files
HAVE_DXF = {'BE018','BE019','BE020','BE021','BE022','BE023','CO006','CO007','CO008','CO009'}
manual = [r for r in manual if r['comp'] in HAVE_DXF]
script = [r for r in script if r['comp'] in HAVE_DXF]

# Match by (comp, pos, hf, parts) — hf=None for CJP is a distinct key
def pair_key(r):
    parts = tuple(sorted((r['p1'], r['p2'])))
    hf = r['hf']  # None → CJP, numeric → fillet
    return (r['comp'], r['pos'], hf, parts)

# Build manual groups
man_groups = defaultdict(list)
for r in manual:
    man_groups[pair_key(r)].append(r['len'])

scr_groups = defaultdict(list)
for r in script:
    scr_groups[pair_key(r)].append(r['len'])

print(f"{'Comp':<8} {'Side':<6} {'hf':>4}  {'Parts':<22}  {'Manual lens':<30}  {'Script lens'}")
print("-" * 110)

def _sort_key(k):
    comp, pos, hf, parts = k
    return (comp, pos, hf if hf is not None else -1, parts)

all_keys = sorted(set(list(man_groups.keys()) + list(scr_groups.keys())),
                  key=_sort_key)

total_len_match = 0
total_len_diff  = 0
total_only_man  = 0
total_only_scr  = 0

for k in all_keys:
    comp, pos, hf, parts = k
    mlens = sorted(man_groups.get(k, []))
    slens = sorted(scr_groups.get(k, []))

    hf_str = 'CJP' if hf is None else str(int(hf)) if isinstance(hf, float) and hf == int(hf) else str(hf)
    if not mlens:
        # script only
        total_only_scr += len(slens)
        tag = "SCRIPT-ONLY"
        print(f"{comp:<8} {pos:<6} {hf_str:>5}  {'/'.join(parts):<22}  {'—':<30}  {slens}  [{tag}]")
    elif not slens:
        # manual only
        total_only_man += len(mlens)
        tag = "MISSED"
        print(f"{comp:<8} {pos:<6} {hf_str:>5}  {'/'.join(parts):<22}  {str(mlens):<30}  —  [{tag}]")
    else:
        # both present — compare lengths
        if mlens == slens:
            total_len_match += len(mlens)
            # only print if you want to see matches too
            # print(f"{comp:<8} {pos:<6} {hf_str:>5}  {'/'.join(parts):<22}  {str(mlens):<30}  {slens}  [OK]")
        else:
            total_len_diff += 1
            tag = "LEN-DIFF"
            print(f"{comp:<8} {pos:<6} {hf_str:>5}  {'/'.join(parts):<22}  {str(mlens):<30}  {slens}  [{tag}]")

print()
print(f"Length-exact matches : {total_len_match}")
print(f"Length mismatches    : {total_len_diff}")
print(f"Script-only keys     : {total_only_scr}")
print(f"Manual-only keys     : {total_only_man}")
