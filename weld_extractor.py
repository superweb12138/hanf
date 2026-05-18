"""
Weld Statistics Extractor
Reads DXF files (converted from DWG) and outputs weld statistics to Excel.

Workflow:
  1. Run convert_dwg_to_dxf.py  (once, converts all DWGs to DXF)
  2. Run explore_dxf.py         (optional, inspect DXF structure)
  3. Run this script             (extracts weld data -> Excel)
"""
import ezdxf
import math
import re
import os
import glob
from collections import defaultdict, Counter

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

# ============================================================
# Configuration
# ============================================================
FOLDER   = r"D:\hanf"
OUTPUT   = os.path.join(FOLDER, "焊缝统计_auto.xlsx")

# Scale: 1 CAD unit = SCALE mm  (confirmed: 44.042 CAD = 440.4 mm → scale=10)
SCALE    = 10.0

# Arrow-tip to Part-line snap tolerance (CAD units)
SNAP_TOL  = 1.5
MAX_HF    = 20    # cap; very large annotations are plate thickness proxies, but hf=16 is valid

# Mark leader-tip to Part-line tolerance for label assignment
LABEL_TIP_TOL = 8.0

# ============================================================
# Geometry helpers
# ============================================================
def dist2d(a, b):
    return math.sqrt((a[0]-b[0])**2 + (a[1]-b[1])**2)

def dist_pt_to_seg(pt, s, e):
    """Perpendicular distance from pt to segment s→e. Returns (dist, t) where t∈[0,1]."""
    dx, dy = e[0]-s[0], e[1]-s[1]
    len_sq = dx*dx + dy*dy
    if len_sq < 1e-12:
        return dist2d(pt, s), 0.0
    t = max(0.0, min(1.0, ((pt[0]-s[0])*dx + (pt[1]-s[1])*dy) / len_sq))
    proj = (s[0]+t*dx, s[1]+t*dy)
    return dist2d(pt, proj), t

def pt_on_seg(pt, s, e, tol):
    d, _ = dist_pt_to_seg(pt, s, e)
    return d <= tol

# ============================================================
# Merge fragmented colinear edges (fix for polyline-drawn parts)
# ============================================================
def _merge_collinear_edges(edges_with_lines, adj_tol):
    """
    Merge fragmented colinear gusset edges that touch the same other_part.
    When a part is drawn as a polyline, its edges are broken into multiple
    short LINE entities.  This merges adjacent, colinear segments that
    touch the same neighbouring part back into a single edge.

    edges_with_lines: list of (length, other_part, gusset_line_dict)
    adj_tol: max endpoint distance to consider two lines touching
    Returns: list of (merged_length, other_part, source_fragments)
    """
    if len(edges_with_lines) <= 1:
        return [(e, op, [g_ln]) for e, op, g_ln in edges_with_lines]

    groups = defaultdict(list)
    for ln_len, op, g_ln in edges_with_lines:
        groups[op].append((ln_len, g_ln))

    merged = []
    for op, items in groups.items():
        if len(items) == 1:
            merged.append((items[0][0], op, [items[0][1]]))
            continue

        n = len(items)
        parent = list(range(n))
        def _find(x):
            while parent[x] != x:
                parent[x] = parent[parent[x]]
                x = parent[x]
            return x
        def _union(a, b):
            parent[_find(a)] = _find(b)

        for i in range(n):
            li = items[i][1]
            for j in range(i + 1, n):
                lj = items[j][1]
                # Endpoint adjacency check
                if (dist2d(li['start'], lj['start']) < adj_tol or
                    dist2d(li['start'], lj['end'])   < adj_tol or
                    dist2d(li['end'],   lj['start']) < adj_tol or
                    dist2d(li['end'],   lj['end'])   < adj_tol):
                    # Colinearity check: avoid merging L-shaped corners
                    dx1 = li['end'][0] - li['start'][0]
                    dy1 = li['end'][1] - li['start'][1]
                    dx2 = lj['end'][0] - lj['start'][0]
                    dy2 = lj['end'][1] - lj['start'][1]
                    len1 = math.hypot(dx1, dy1)
                    len2 = math.hypot(dx2, dy2)
                    if len1 > 1e-9 and len2 > 1e-9:
                        cos_a = abs(dx1 * dx2 + dy1 * dy2) / (len1 * len2)
                        if cos_a > 0.985:   # cos(10°) — same line direction
                            _union(i, j)

        comps = defaultdict(list)
        for i in range(n):
            comps[_find(i)].append(items[i])

        for comp_items in comps.values():
            total_len = sum(it[0] for it in comp_items)
            source_fragments = [it[1] for it in comp_items]
            merged.append((total_len, op, source_fragments))

    return merged

# ============================================================
# WeldMark parsing
# ============================================================
def parse_weldmark(blk):
    """
    Extract weld data from a WeldMark block definition.
    Returns dict or None.

    Strategy:
      - Collect all line endpoints; dangling endpoints (count==1) are candidates.
      - The horizontal reference shelf has y ≈ constant and len > 3 units.
      - Arrow tip = dangling endpoint whose y differs from the reference shelf y.
      - Sizes come from numeric TEXT entities; y-position relative to shelf →
        above (other-side weld) or below (arrow-side weld).
    """
    lines_raw = []
    texts = []
    for e in blk:
        t = e.dxftype()
        if t == 'LINE':
            s  = (round(e.dxf.start.x, 4), round(e.dxf.start.y, 4))
            ep = (round(e.dxf.end.x,   4), round(e.dxf.end.y,   4))
            ln = dist2d(s, ep)
            if ln > 0.01:
                lines_raw.append((s, ep, ln))
        elif t == 'TEXT':
            try:
                txt = e.dxf.text.strip()
                pos = (e.dxf.insert.x, e.dxf.insert.y)
                if txt:
                    texts.append((txt, pos))
            except:
                pass
        elif t == 'MTEXT':
            try:
                txt = e.text.strip()
                pos = (e.dxf.insert.x, e.dxf.insert.y)
                if txt:
                    texts.append((txt, pos))
            except:
                pass

    if not lines_raw:
        return None

    arclist = []
    for e in blk:
        if e.dxftype() == "ARC":
            c = (round(e.dxf.center.x, 4), round(e.dxf.center.y, 4))
            r = round(e.dxf.radius, 4)
            arclist.append((c, r))

    # Dangling endpoints
    ep_count = Counter()
    for s, ep, _ in lines_raw:
        ep_count[s]  += 1
        ep_count[ep] += 1
    dangling = {pt for pt, c in ep_count.items() if c == 1}

    # Reference shelf: longest horizontal line
    horiz = [(s, ep, ln) for s, ep, ln in lines_raw
             if abs(s[1]-ep[1]) < 0.05*ln and ln > 3]
    if not horiz:
        return None
    ref_s, ref_e, _ = max(horiz, key=lambda x: x[2])
    ref_y = (ref_s[1] + ref_e[1]) / 2.0

    _cc = Counter()
    for c, r in arclist:
        if 1.0 <= r <= 2.5 and abs(c[1] - ref_y) < 1.0:
            _cc[c] += 1
    has_circle = any(cnt >= 2 for cnt in _cc.values())

    # Arrow tip candidates: dangling points NOT on the shelf y-level
    candidates = [pt for pt in dangling if abs(pt[1] - ref_y) > 0.5]
    if not candidates:
        return None
    arrow_tip = max(candidates, key=lambda pt: abs(pt[1] - ref_y))

    # Parse weld sizes from text
    size_above = None   # other-side (above shelf)
    size_below = None   # arrow-side (below shelf)
    groove_above = False  # True if above side is CJP/groove (hf=0)
    groove_below = False  # True if below side is CJP/groove (hf=0)
    annotation = ""
    for txt, pos in texts:
        m = re.match(r'^(\d+(?:\.\d+)?)(?:/(\d+(?:\.\d+)?))?$', txt)
        if m:
            sz = float(m.group(1))
            if pos[1] >= ref_y:
                size_above = sz
            else:
                size_below = sz
        elif 'CJP' in txt.upper():
            # CJP (complete joint penetration) = groove weld; mark the side it appears on
            if pos[1] >= ref_y:
                groove_above = True
            else:
                groove_below = True
        elif any(kw in txt.upper() for kw in ['SIDE', '围', '全', 'ALL']):
            annotation = txt
    # TYP / TYP. = typical weld (applies to multiple symmetric instances)
    is_typ = any('TYP' in txt.upper() for txt, _ in texts)
    # Groove/CJP: the groove side keeps size=None (full penetration, no leg size).
    # If a valid fillet < MAX_HF is already present on the same side, the groove
    # annotation belongs to a different weld path — keep the fillet size.
    if groove_above and (size_above is None or size_above >= MAX_HF):
        size_above = None   # CJP / groove → no fillet size
    if groove_below and (size_below is None or size_below >= MAX_HF):
        size_below = None   # CJP / groove → no fillet size
    # When CJP/groove is on one side and a valid fillet size is also on
    # that side (but the opposite side has no size), the fillet is the
    # paired fillet on the opposite face — move it across.  Keep the
    # groove flag so the fillet bypasses the thickness-correction step.
    if groove_above and size_above is not None and size_above <= MAX_HF and size_below is None:
        size_below = size_above; size_above = None
    if groove_below and size_below is not None and size_below <= MAX_HF and size_above is None:
        size_above = size_below; size_below = None
    if groove_above and size_above is not None and size_above <= MAX_HF:
        groove_above = False  # valid fillet present, groove is separate notation
    if groove_below and size_below is not None and size_below <= MAX_HF:
        groove_below = False
    # Numbers > MAX_HF are plate-thickness annotations, treat as no fillet
    if size_above is not None and size_above > MAX_HF:
        size_above = None
    if size_below is not None and size_below > MAX_HF:
        size_below = None

    # Detect triangle symbols above/below shelf
    # Triangle lines are short (≈3-6 units); a vertex above or below the shelf confirms that side
    has_above = any(
        max(s[1], ep[1]) > ref_y + 0.5
        for s, ep, ln in lines_raw if ln < 7
    )
    has_below = any(
        min(s[1], ep[1]) < ref_y - 0.5
        for s, ep, ln in lines_raw if ln < 7
    )

    if has_circle:
        has_above = True
        if size_above is None and size_below is not None:
            size_above = size_below
        elif size_below is None and size_above is not None:
            size_below = size_above

    # Leader line endpoint list (for multi-segment leaders)
    # The longest non-horizontal line is usually the main leader
    non_horiz = [(s, ep, ln) for s, ep, ln in lines_raw
                 if not (abs(s[1]-ep[1]) < 0.05*ln and ln > 3)]

    return {
        'arrow_tip':   arrow_tip,
        'size_above':  size_above,
        'size_below':  size_below,
        'has_above':   has_above,
        'has_below':   has_below,
        'annotation':  annotation,
        'groove_above': groove_above,
        'groove_below': groove_below,
        'is_typ':      is_typ,
        'has_circle':  has_circle,
        'ref_y':       ref_y,
        'texts':       texts,
    }

# ============================================================
# Part geometry
# ============================================================
def get_part_lines(blk):
    """Return list of {start, end, length} dicts for all lines in a Part block."""
    lines = []
    for e in blk:
        if e.dxftype() == 'LINE':
            s  = (e.dxf.start.x, e.dxf.start.y)
            ep = (e.dxf.end.x,   e.dxf.end.y)
            ln = dist2d(s, ep)
            if ln > 0.5:
                lines.append({'start': s, 'end': ep, 'length': ln})
    return lines

def part_centroid(lines):
    if not lines:
        return (0.0, 0.0)
    xs = [(l['start'][0]+l['end'][0])/2 for l in lines]
    ys = [(l['start'][1]+l['end'][1])/2 for l in lines]
    return (sum(xs)/len(xs), sum(ys)/len(ys))

# ============================================================
# Find part labels (text that looks like a part number)
# ============================================================
PART_RE = re.compile(r'^[pP]\d+$|^[A-Z]{2,3}\d+$|^\d{3,}$')

def find_all_labels(doc):
    """
    Scan Mark blocks for text matching part-number patterns.
    Extracts leader_tip = farthest line endpoint from the text position,
    which is the point where the leader arrow touches the labelled part.
    """
    labels = []
    for blk in doc.blocks:
        blk_name = blk.name
        if not blk_name.startswith('Mark'):
            continue
        txt_pos = None
        texts   = []
        lines   = []
        for e in blk:
            if e.dxftype() == 'TEXT':
                try:
                    t = e.dxf.text.strip()
                    if t:
                        texts.append(t)
                    if txt_pos is None:
                        txt_pos = (e.dxf.insert.x, e.dxf.insert.y)
                except:
                    pass
            elif e.dxftype() == 'MTEXT':
                try:
                    t = e.text.strip()
                    if t:
                        texts.append(t)
                    if txt_pos is None:
                        txt_pos = (e.dxf.insert.x, e.dxf.insert.y)
                except:
                    pass
            elif e.dxftype() == 'LINE':
                try:
                    lines.append(((e.dxf.start.x, e.dxf.start.y),
                                  (e.dxf.end.x,   e.dxf.end.y)))
                except:
                    pass
        label = next((t for t in texts if PART_RE.match(t)), None)
        if not label or not txt_pos:
            continue
        if lines:
            all_pts    = [p for seg in lines for p in seg]
            leader_tip = max(all_pts, key=lambda p: dist2d(p, txt_pos))
        else:
            leader_tip = txt_pos
        labels.append({'label': label, 'pos': txt_pos,
                       'leader_tip': leader_tip, 'block': blk_name})
    return labels


def assign_labels_by_leader_tip(all_labels, part_lines_map):
    """
    Match each Mark block's leader tip to the nearest Part line in the same view.
    Uses centroid distance as tiebreaker when line distances are essentially equal,
    ensuring adjacent parts sharing a face line are distinguished correctly.
    The same label string can be assigned to one Part per view (multi-view drawings).
    Returns: {part_name -> label_string}
    """
    part_number_map = {}
    for lbl in all_labels:
        m = re.search(r' - (\d+)$', lbl['block'])
        if not m:
            continue
        view_id    = m.group(1)
        tip        = lbl['leader_tip']
        view_parts = part_lines_map.get(view_id, {})
        best_part  = None
        best_score = (LABEL_TIP_TOL, 1e18)   # (line_dist, centroid_dist)
        for pname, lines in view_parts.items():
            line_d = LABEL_TIP_TOL
            for ln in lines:
                d, _ = dist_pt_to_seg(tip, ln['start'], ln['end'])
                d    = min(d, dist2d(tip, ln['start']), dist2d(tip, ln['end']))
                if d < line_d:
                    line_d = d
            if line_d < LABEL_TIP_TOL:
                c  = part_centroid(lines) if lines else tip
                cd = dist2d(tip, c)
                score = (line_d, cd)
                if score < best_score:
                    best_score = score
                    best_part  = pname
        if best_part:
            part_number_map[best_part] = lbl['label']
    return part_number_map

# ============================================================
# Spatial matching
# ============================================================
def find_parts_at_point(arrow_tip, view_part_lines, tol):
    """
    Return list of match dicts where the arrow_tip lies on or near a Part line.
    Each dict: {'part', 'how' (endpoint|interior), 'line', 'ep_dist'|'int_dist'}
    Per part: keep the closest-endpoint match, or shortest interior match.
    """
    matches = []
    for part_name, lines in view_part_lines.items():
        best_ep  = None   # (line_dict, ep_dist)
        best_int = None   # (line_dict, int_dist)
        for ln in lines:
            d_start = dist2d(arrow_tip, ln['start'])
            d_end   = dist2d(arrow_tip, ln['end'])
            ep_d    = min(d_start, d_end)
            if ep_d <= tol:
                if best_ep is None or ep_d < best_ep[1]:
                    best_ep = (ln, ep_d)
            else:
                d_int, _ = dist_pt_to_seg(arrow_tip, ln['start'], ln['end'])
                if d_int <= tol:
                    # Keep the shortest interior line (= weld seam, not main member)
                    if best_int is None or ln['length'] < best_int[0]['length']:
                        best_int = (ln, d_int)
        if best_ep is not None:
            matches.append({'part': part_name, 'how': 'endpoint',
                            'line': best_ep[0],  'ep_dist': best_ep[1]})
        elif best_int is not None:
            matches.append({'part': part_name, 'how': 'interior',
                            'line': best_int[0], 'int_dist': best_int[1]})
    return matches

# ============================================================
# Determine weld length for a given arrow tip + matched parts
# ============================================================
def choose_weld_line(arrow_tip, matches):
    """
    Given match dicts at the arrow tip, choose the weld line (part + line).
    Tier 1: exact endpoint (ep_dist < 0.3) with line >= MIN_LINE
    Tier 2: any endpoint with line >= MIN_LINE
    Tier 3: interior matches (shortest first)
    Tier 4: any match (last resort)
    Short cross-section stubs (< MIN_LINE) are skipped in endpoint tiers so
    that an interior match on the actual weld seam takes precedence.
    Returns (part, line, match_how) where match_how is 'endpoint' or 'interior'.
    """
    MIN_LINE = 2.0   # CAD units; flange/web thickness stubs are below this

    if not matches:
        return None, None, None

    # Tier 1: exact endpoint, substantive line
    t1 = [m for m in matches
          if m['how'] == 'endpoint' and m['ep_dist'] < 0.3
          and m['line']['length'] >= MIN_LINE]
    if t1:
        best = min(t1, key=lambda m: m['line']['length'])
        return best['part'], best['line'], 'endpoint'

    # Tier 2: any endpoint, substantive line
    t2 = [m for m in matches
          if m['how'] == 'endpoint' and m['line']['length'] >= MIN_LINE]
    if t2:
        best = min(t2, key=lambda m: m['line']['length'])
        return best['part'], best['line'], 'endpoint'

    # Tier 3: interior matches (prefer shortest = weld seam, not main member)
    t3 = [m for m in matches if m['how'] == 'interior']
    t3_good = [m for m in t3 if m['line']['length'] >= MIN_LINE]
    pool = t3_good if t3_good else t3
    if pool:
        best = min(pool, key=lambda m: m['line']['length'])
        return best['part'], best['line'], 'interior'

    # Tier 4: last resort (any match including short stubs)
    best = min(matches, key=lambda m: m['line']['length'])
    return best['part'], best['line'], 'interior'

# ============================================================
# Standard fillet size table  (Sub-rule 3: plate/web thickness → hf)
# ============================================================
_HF_FROM_T = {6:5, 7:5, 8:6, 9:6, 10:7, 11:8, 12:8, 14:10, 16:10, 18:12, 20:12}

def hf_from_thickness(t):
    """Return standard min fillet size for a given plate/web thickness (mm)."""
    t = int(round(t))
    if t in _HF_FROM_T:
        return _HF_FROM_T[t]
    if t <= 6:  return 5
    if t <= 12: return int(round(t * 0.67))
    return 10

# ============================================================
# BOM parser  (Unknown block part schedule)
# ============================================================
def parse_bom(doc, comp):
    """
    Parse the part schedule (BOM) from the Unknown block that contains
    part mark + PLt×W / HWd×b×tw×tf entries.

    Returns:
      part_dims  : {label -> {'thick': t, 'width': w, 'bom_len': l, 'qty': q}}
      comp_dims  : {'depth': d, 'flange_w': b, 'web_t': tw, 'flange_t': tf} or {}
    """
    part_dims = {}
    comp_dims = {}

    for blk in doc.blocks:
        # Global Unknown blocks only (no " - XXXX" suffix)
        if not (blk.name.startswith('Unknown') and ' - ' not in blk.name):
            continue

        # Collect all TEXT/MTEXT with position
        raw = []
        for e in blk:
            if e.dxftype() not in ('TEXT', 'MTEXT'):
                continue
            try:
                txt = (e.dxf.text if e.dxftype() == 'TEXT' else e.text).strip()
                x = round(e.dxf.insert.x, 0)
                y = round(e.dxf.insert.y, 1)
                if txt:
                    raw.append((y, x, txt))
            except:
                pass

        # Group into rows by y (bucket 4-unit bands)
        rows = defaultdict(dict)
        for y, x, txt in raw:
            rows[round(y / 4) * 4][x] = txt

        found_any = False
        for yk in sorted(rows, reverse=True):
            rowvals = rows[yk]
            # Sort by x-coordinate so columns read left→right:
            #  [drawing#] [seq] [qty] [mark] [spec] [grade] [len] [note] [weight]
            vals_sorted = sorted(rowvals.items())
            vals = [txt for _, txt in vals_sorted]
            mark  = next((v for v in vals if re.match(r'^p\d+$', v) or v == comp), None)
            spec  = next((v for v in vals if re.search(r'(?:PL|HW|HN|HM)\d+[xX]', v, re.I)), None)
            if not (mark and spec):
                continue
            found_any = True
            # Parse plate spec PLt×W or PLt×W×L
            pm = re.match(r'PL(\d+(?:\.\d+)?)[xX×](\d+(?:\.\d+)?)', spec, re.I)
            # Parse H-section HWd×b×tw×tf
            hm = re.match(r'H[WNMQwq](\d+(?:\.\d+)?)[xX×](\d+(?:\.\d+)?)[xX×](\d+(?:\.\d+)?)[xX×](\d+(?:\.\d+)?)', spec, re.I)
            # BOM length column (largest number > 50 in the row, not the spec itself)
            nums = []
            for v in vals:
                if v == spec: continue
                try:
                    fv = float(v)
                    if fv > 50:
                        nums.append(fv)
                except:
                    pass
            bom_len = max(nums) if nums else None
            # Qty: second 1-2 digit number (first is seq number, see column order above)
            small_nums = [int(v) for v in vals if re.match(r'^\d{1,2}$', v)]
            qty = small_nums[1] if len(small_nums) >= 2 else (small_nums[0] if small_nums else 1)

            if pm:
                t, w = float(pm.group(1)), float(pm.group(2))
                # Filter: if the found "length" is unreasonably large compared
                # to the plate width (e.g. weight column misread as length),
                # discard it.  Typical plate aspect ratio L/W <= 4.
                if bom_len and w > 0 and bom_len > w * 4:
                    bom_len = None
                part_dims[mark] = {'thick': t, 'width': w, 'bom_len': bom_len, 'qty': qty}
            elif hm and mark == comp:
                d, b, tw, tf = (float(hm.group(i)) for i in (1, 2, 3, 4))
                comp_dims = {'depth': d, 'flange_w': b, 'web_t': tw, 'flange_t': tf}
                part_dims[mark] = {'thick': tf, 'width': b, 'bom_len': bom_len, 'qty': qty}

        if found_any:
            break   # use first BOM block found

    return part_dims, comp_dims

# ============================================================
# Main per-file extraction
# ============================================================
def extract_welds(dxf_path):
    comp_m = re.search(r'-(BE\d+|CO\d+)_', os.path.basename(dxf_path), re.I)
    comp   = comp_m.group(1).upper() if comp_m else os.path.splitext(os.path.basename(dxf_path))[0]

    print(f"\n{'='*60}\n{os.path.basename(dxf_path)}  [{comp}]")

    doc = ezdxf.readfile(dxf_path)

    # Parse BOM for part dimensions and comp section properties
    part_dims, comp_dims = parse_bom(doc, comp)
    comp_web_t    = comp_dims.get('web_t',    None)   # e.g. 9  for HW250×250×9×14
    comp_flange_t = comp_dims.get('flange_t', None)   # e.g. 14
    print(f"  BOM parts: {list(part_dims.keys())}")
    if comp_dims:
        print(f"  Comp section: {comp_dims}")

    def _correct_hf(sz, lbl_a, lbl_b):
        """Sub-rule 3: replace plate/web-thickness annotation with standard fillet size.
        Only applied for sz <= 12 to preserve valid large fillets (e.g. hf=16 for CO009)."""
        if sz is None or sz > 12:
            return sz
        if comp_web_t and abs(sz - comp_web_t) < 0.5:
            return hf_from_thickness(comp_web_t)
        if comp_flange_t and abs(sz - comp_flange_t) < 0.5:
            return hf_from_thickness(comp_flange_t)
        for lbl in (lbl_a, lbl_b):
            if lbl != comp and lbl in part_dims:
                t = part_dims[lbl]['thick']
                if abs(sz - t) < 0.5 and sz > 8:
                    return hf_from_thickness(t)
        return sz

    def _correct_hf_3s(sz, lbl_gusset):
        """hf correction for 3-SIDES: check against gusset thickness and comp web.
        Only applied for sz <= 12 to preserve valid large fillets."""
        if sz is None or sz > 12:
            return sz
        if comp_web_t and abs(sz - comp_web_t) < 0.5:
            return hf_from_thickness(comp_web_t)
        if lbl_gusset in part_dims:
            t = part_dims[lbl_gusset]['thick']
            if abs(sz - t) < 0.5 and sz > 8:
                return hf_from_thickness(t)
        return sz

    # Group WeldMark and Part blocks by view ID (suffix " - XXXX")
    wm_by_view   = defaultdict(list)   # view_id -> [(name, blk)]
    part_by_view = defaultdict(list)   # view_id -> [(name, blk)]

    for blk in doc.blocks:
        blk_name = blk.name
        m = re.search(r' - (\d+)$', blk_name)
        if not m:
            continue
        view_id = m.group(1)
        if blk_name.startswith('WeldMark'):
            wm_by_view[view_id].append((blk_name, blk))
        elif blk_name.startswith('Part'):
            part_by_view[view_id].append((blk_name, blk))

    print(f"  Views with WeldMarks : {sorted(wm_by_view)}")
    print(f"  Views with Parts     : {sorted(part_by_view)}")

    # Build part geometry maps
    part_lines_map = {}    # view_id -> {part_name: [lines]}

    for view_id, parts in part_by_view.items():
        part_lines_map[view_id] = {}
        for pname, pblk in parts:
            lines = get_part_lines(pblk)
            part_lines_map[view_id][pname] = lines

    # Assign part labels via Mark block leader tips
    all_labels      = find_all_labels(doc)
    part_number_map = assign_labels_by_leader_tip(all_labels, part_lines_map)
    print(f"  Part label candidates: {[x['label'] for x in all_labels]}")
    print(f"  Part→label map : {part_number_map}")

    # Infer dimensions for non-BOM parts by geometry analysis.
    # For CO components, many stiffener parts (p183, p197, etc.) are not
    # listed in the BOM but we need width/length for the CO-fallback and
    # thickness for hf correction.
    _inferred = {}
    for _pn, _lbl in part_number_map.items():
        if _lbl == comp or _lbl in part_dims:
            continue
        # Collect lines for this label across all views
        _all_lns = []
        for _vid, _pmap in part_lines_map.items():
            if _pn in _pmap:
                _all_lns.extend(_pmap[_pn])
        if not _all_lns:
            continue
        # Bounding-box based dimension estimate
        _xs = [p[0] for ln in _all_lns for p in (ln['start'], ln['end'])]
        _ys = [p[1] for ln in _all_lns for p in (ln['start'], ln['end'])]
        _w = max(_xs) - min(_xs)
        _h = max(_ys) - min(_ys)
        _w_mm = round(_w * SCALE, 1)
        _h_mm = round(_h * SCALE, 1)
        _bw = min(_w_mm, _h_mm)
        _bl = max(_w_mm, _h_mm)
        # Thickness: use comp_web_t if available, else default 12mm
        _t = comp_web_t if comp_web_t else 12
        # qty stays 1 — TYP multiplier uses main_view count, not all-view instances
        _inferred[_lbl] = {'thick': _t, 'width': _bw, 'bom_len': _bl, 'qty': 1}
    if _inferred:
        _inf_strs = []
        for _lbl, _dim in _inferred.items():
            _inf_strs.append('%s:w=%s L=%s qty=%s' % (_lbl, _dim['width'], _dim['bom_len'], _dim['qty']))
        print('  [infer-dims] %s' % _inf_strs)
    # Merge inferred into part_dims (inferred don't overwrite existing BOM data)
    for _lbl, _dim in _inferred.items():
        if _lbl not in part_dims:
            part_dims[_lbl] = _dim

    # Extract welds
    results = []
    skipped = []

    # Determine main view: the view with the most Part-block instances.
    # Used to resolve TYP (typical) multipliers — TYP welds appear once on the
    # drawing but represent every instance of that part in the main assembly view.
    from collections import Counter as _Ctr
    _view_cnt = _Ctr(k.split(' - ')[-1] for k in part_number_map)
    main_view_id = _view_cnt.most_common(1)[0][0] if _view_cnt else ''

    # Cross-view dedup for 3-SIDES edges: same gusset + same other part
    # + same geo length in a different view → same physical edge.
    cross_view_seen = {}  # (gusset_label, other_label, geo_mm) → view_id

    for view_id, weldmarks in wm_by_view.items():
        view_parts = part_lines_map.get(view_id, {})
        if not view_parts:
            print(f"  View {view_id}: no Part blocks found, skipping {len(weldmarks)} WeldMark(s)")
            continue

        for wm_name, wm_blk in weldmarks:
            parsed = parse_weldmark(wm_blk)
            if not parsed:
                skipped.append((wm_name, "parse failed"))
                continue

            arrow   = parsed['arrow_tip']
            matches = find_parts_at_point(arrow, view_parts, SNAP_TOL)

            if not matches:
                skipped.append((wm_name, f"no Part at arrow_tip {arrow}"))
                continue

            # '3 SIDES' / '2 SIDES' / '围' / '全' all indicate a perimeter gusset weld
            # where edges of the attachment plate must be enumerated.
            # (Note: "2 SIDES" can mean double-sided fillet, not necessarily
            # 2 physical edges — so we always assume the gusset's welded edges
            # number is the plate's perimeter count minus free edges.)
            is_three_sides = any(kw in parsed['annotation'].upper()
                                 for kw in ['SIDE', '围', '全'])
            is_circle_wm = parsed.get('has_circle', False)
            _use_largest_gusset = False
            if is_circle_wm and not is_three_sides:
                comp_part_names_x = {pn for pn, lbl in part_number_map.items() if lbl == comp}
                non_comp_matches_x = [m for m in matches if m['part'] not in comp_part_names_x]
                if non_comp_matches_x:
                    is_three_sides = True
                    _use_largest_gusset = True
            expected_edges = 10 if is_circle_wm else 3

            if is_three_sides:
                # Gusset = the smallest-line NON-COMP Part at the arrow.
                # The comp (main member) is never the gusset plate.
                comp_part_names = {pn for pn, lbl in part_number_map.items() if lbl == comp}
                non_comp_matches = [m for m in matches if m['part'] not in comp_part_names]
                gusset_pool  = non_comp_matches if non_comp_matches else matches
                if _use_largest_gusset:
                    _max_gust_len = max(m['line']['length'] for m in gusset_pool)
                    gusset_names = list(dict.fromkeys(
                        m['part'] for m in gusset_pool
                        if m['line']['length'] >= _max_gust_len * 0.95
                    ))
                else:
                    _min_gust_len = min(m['line']['length'] for m in gusset_pool)
                    gusset_names = list(dict.fromkeys(
                        m['part'] for m in gusset_pool
                        if m['line']['length'] <= _min_gust_len * 1.05
                    ))
                # For multi-gusset (same label, multiple instances) collect all
                # edges; for single gusset the loop runs once (no change)
                gusset_name     = gusset_names[0]   # primary (used for label/print)
                gusset_blk_set  = set(gusset_names)  # skip ALL gussets as neighbors
                ADJ_TOL         = SNAP_TOL + 0.5
                MIN_EDGE        = 1.5  # CAD units (<15mm = degenerate stub)

                _synth = _use_largest_gusset and bool(comp_dims.get('flange_w'))
                if _synth:
                    _, _wl, _ = choose_weld_line(arrow, matches)
                    _mid_cad = _wl['length'] if _wl else 0
                    _fw_cad = comp_dims['flange_w'] / SCALE
                    if _mid_cad > 0 and _fw_cad > 0:
                        _comp_blk = next((pn for pn, lbl in part_number_map.items()
                                         if lbl == comp), gusset_name)
                        _dum_fw = {'start':(0,0),'end':(_fw_cad,0),'length':_fw_cad}
                        _dum_md = {'start':(0,0),'end':(_mid_cad,0),'length':_mid_cad}
                        _se = [(_fw_cad,_comp_blk,[_dum_fw]),(_mid_cad,_comp_blk,[_dum_md]),(_fw_cad,_comp_blk,[_dum_fw])]
                        weld_edges_by_gusset = {}
                        for _gn in gusset_names:
                            weld_edges_by_gusset[_gn] = _se
                    else:
                        skipped.append((wm_name, 'CIRCLE: no weld line'))
                        continue
                if not _synth:

                # Collect edges per gusset block (each gusset processed independently)
                # to support multi-instance assemblies (e.g. haunches on both flanges).
                # Build pass-through for unlabeled parts: find which labeled
                # part each unlabeled part is adjacent to (e.g. a small filler
                # plate sandwiched between the gusset and the main part).
                    unlabeled_passthru = {}
                    _unlabeled = {pn for pn in view_parts
                                  if pn not in part_number_map and pn not in gusset_blk_set}
                    for _up in _unlabeled:
                        _ulns = view_parts[_up]
                        _best_lbl = None; _best_d = ADJ_TOL
                        for _lpn, _llns in view_parts.items():
                            if _lpn in gusset_blk_set or _lpn not in part_number_map:
                                continue
                            for _uln in _ulns:
                                for _lln in _llns:
                                    _d = min(
                                        math.hypot(_uln['start'][0]-_lln['start'][0], _uln['start'][1]-_lln['start'][1]),
                                        math.hypot(_uln['start'][0]-_lln['end'][0], _uln['start'][1]-_lln['end'][1]),
                                        math.hypot(_uln['end'][0]-_lln['start'][0], _uln['end'][1]-_lln['start'][1]),
                                        math.hypot(_uln['end'][0]-_lln['end'][0], _uln['end'][1]-_lln['end'][1]),
                                    )
                                    if _d < _best_d:
                                        _best_d = _d; _best_lbl = part_number_map.get(_lpn, comp)
                        if _best_lbl:
                            unlabeled_passthru[_up] = _best_lbl
                    # Prevent passthru from mapping unlabeled parts to the
                    # gusset's own label (creates self-reference that gets
                    # skipped — the edge should go to comp instead).
                    _lbl_g = part_number_map.get(gusset_name, comp)
                    if _lbl_g != comp:
                        for _up in list(unlabeled_passthru):
                            if unlabeled_passthru[_up] == _lbl_g:
                                del unlabeled_passthru[_up]
                    if unlabeled_passthru:
                        print(f"    [unlabeled→label] {unlabeled_passthru}")

                    weld_edges_by_gusset = {}
                    for _gn in gusset_names:
                        _edges = []
                        for g_ln in view_parts.get(_gn, []):
                            if g_ln['length'] < MIN_EDGE:
                                continue
                            p_s = None; pd_s = ADJ_TOL
                            p_e = None; pd_e = ADJ_TOL
                            for pname, plines in view_parts.items():
                                if pname in gusset_blk_set:
                                    continue
                                for ln in plines:
                                    d1, _ = dist_pt_to_seg(g_ln['start'], ln['start'], ln['end'])
                                    d2, _ = dist_pt_to_seg(g_ln['end'],   ln['start'], ln['end'])
                                    if d1 <= pd_s: pd_s = d1; p_s = pname
                                    if d2 <= pd_e: pd_e = d2; p_e = pname
                            # For unlabeled neighbours, scan the gusset edge directly
                            # against all labelled non-comp parts to find the real
                            # neighbour.
                            _orig_ps, _orig_pe = p_s, p_e
                            _lbl_gn = part_number_map.get(_gn, comp)
                            if p_s and p_s not in part_number_map:
                                    _ps_is_comp = False
                                    if comp_dims:
                                        _cl = {round(comp_dims.get('flange_w',0)), round(comp_dims.get('depth',0))}
                                        _cl.discard(0)
                                        for _tln in view_parts.get(p_s, []):
                                            if round(_tln['length'] * SCALE) in _cl:
                                                _ps_is_comp = True; break
                                    if not _ps_is_comp:
                                        _best_r = None; _best_rd = ADJ_TOL * 5
                                        for _pn2, _pln2 in view_parts.items():
                                            if _pn2 in gusset_blk_set:
                                                continue
                                            _lbl2 = part_number_map.get(_pn2, comp)
                                            if _lbl2 == comp or _lbl2 == _lbl_gn:
                                                continue
                                            for _ln2 in _pln2:
                                                _d = min(
                                                    math.hypot(g_ln['start'][0]-_ln2['start'][0], g_ln['start'][1]-_ln2['start'][1]),
                                                    math.hypot(g_ln['start'][0]-_ln2['end'][0],   g_ln['start'][1]-_ln2['end'][1]),
                                                    math.hypot(g_ln['end'][0]  -_ln2['start'][0], g_ln['end'][1]  -_ln2['start'][1]),
                                                    math.hypot(g_ln['end'][0]  -_ln2['end'][0],   g_ln['end'][1]  -_ln2['end'][1]))
                                                if _d < _best_rd:
                                                    _best_rd = _d
                                                    for _lpn2, _lbl2b in part_number_map.items():
                                                        if _lbl2b == _lbl2 and _lpn2.split(' - ')[-1] == view_id:
                                                            _best_r = _lpn2; break
                                        if _best_r: p_s = _best_r
                            if p_e and p_e not in part_number_map:
                                    _pe_is_comp = False
                                    if comp_dims:
                                        _cl2 = {round(comp_dims.get('flange_w',0)), round(comp_dims.get('depth',0))}
                                        _cl2.discard(0)
                                        for _tln2 in view_parts.get(p_e, []):
                                            if round(_tln2['length'] * SCALE) in _cl2:
                                                _pe_is_comp = True; break
                                    if not _pe_is_comp:
                                        _best_r = None; _best_rd = ADJ_TOL * 5
                                        for _pn2, _pln2 in view_parts.items():
                                            if _pn2 in gusset_blk_set:
                                                continue
                                            _lbl2 = part_number_map.get(_pn2, comp)
                                            if _lbl2 == comp or _lbl2 == _lbl_gn:
                                                continue
                                            for _ln2 in _pln2:
                                                _d = min(
                                                    math.hypot(g_ln['start'][0]-_ln2['start'][0], g_ln['start'][1]-_ln2['start'][1]),
                                                    math.hypot(g_ln['start'][0]-_ln2['end'][0],   g_ln['start'][1]-_ln2['end'][1]),
                                                    math.hypot(g_ln['end'][0]  -_ln2['start'][0], g_ln['end'][1]  -_ln2['start'][1]),
                                                    math.hypot(g_ln['end'][0]  -_ln2['end'][0],   g_ln['end'][1]  -_ln2['end'][1]))
                                                if _d < _best_rd:
                                                    _best_rd = _d
                                                    for _lpn2, _lbl2b in part_number_map.items():
                                                        if _lbl2b == _lbl2 and _lpn2.split(' - ')[-1] == view_id:
                                                            _best_r = _lpn2; break
                                        if _best_r: p_e = _best_r
                            if p_s and p_e:
                                g_ln['nb_start'] = _orig_ps
                                g_ln['nb_end']   = _orig_pe
                                if p_s == p_e: _edges.append((g_ln['length'], p_s, g_ln))
                            elif p_s:
                                g_ln['nb_start'] = _orig_ps
                                g_ln['nb_end']   = None
                                if part_number_map.get(p_s, comp) != comp or p_s in unlabeled_passthru:
                                    _edges.append((g_ln['length'], p_s, g_ln))
                            elif p_e:
                                g_ln['nb_start'] = None
                                g_ln['nb_end']   = _orig_pe
                                if part_number_map.get(p_e, comp) != comp or p_e in unlabeled_passthru:
                                    _edges.append((g_ln['length'], p_e, g_ln))
                        # Connected-part enumeration when gusset IS the comp body.
                        if part_number_map.get(_gn, comp) == comp and _gn == gusset_name:
                            for _cpn, _cplns in view_parts.items():
                                if _cpn in gusset_blk_set:
                                    continue
                                _cplbl = part_number_map.get(_cpn, comp)
                                if _cplbl == comp:
                                    continue
                                _touches = False
                                for _cln in _cplns:
                                    if _cln['length'] < MIN_EDGE:
                                        continue
                                    for _gnb in gusset_names:
                                        for _gln in view_parts.get(_gnb, []):
                                            _d1, _ = dist_pt_to_seg(_cln['start'], _gln['start'], _gln['end'])
                                            _d2, _ = dist_pt_to_seg(_cln['end'],   _gln['start'], _gln['end'])
                                            if min(_d1, _d2) <= ADJ_TOL:
                                                _touches = True; break
                                        if _touches: break
                                    if _touches: break
                                if not _touches:
                                    continue
                                for _cln in _cplns:
                                    if _cln['length'] < MIN_EDGE:
                                        continue
                                    _cp_s = None; _cps_d = ADJ_TOL
                                    _cp_e = None; _cpe_d = ADJ_TOL
                                    for _opn, _olns in view_parts.items():
                                        if _opn == _cpn:
                                            continue
                                        for _oln in _olns:
                                            _d1, _ = dist_pt_to_seg(_cln['start'], _oln['start'], _oln['end'])
                                            _d2, _ = dist_pt_to_seg(_cln['end'],   _oln['start'], _oln['end'])
                                            if _d1 <= _cps_d: _cps_d = _d1; _cp_s = _opn
                                            if _d2 <= _cpe_d: _cpe_d = _d2; _cp_e = _opn
                                    if not _cp_s and not _cp_e:
                                        continue
                                    _cln['nb_start'] = _cp_s
                                    _cln['nb_end']   = _cp_e
                                    _s_gust = _cp_s in gusset_blk_set
                                    _e_gust = _cp_e in gusset_blk_set
                                    if _cpn not in weld_edges_by_gusset:
                                        weld_edges_by_gusset[_cpn] = []
                                    _ce = weld_edges_by_gusset[_cpn]
                                    if _cp_s and _cp_e and _cp_s == _cp_e:
                                        _ce.append((_cln['length'], _cp_s, [_cln]))
                                    elif _cp_s and _cp_e:
                                        if _s_gust and not _e_gust:
                                            _ce.append((_cln['length'], _cp_s, [_cln]))
                                        elif _e_gust and not _s_gust:
                                            _ce.append((_cln['length'], _cp_e, [_cln]))
                                    elif _cp_s and _s_gust:
                                        _ce.append((_cln['length'], _cp_s, [_cln]))
                                    elif _cp_e and _e_gust:
                                        _ce.append((_cln['length'], _cp_e, [_cln]))
                        # Dedup exact duplicates before merging (same start+end)
                        _dedup = []
                        _seen_pts = set()
                        for _e in _edges:
                            _pts = (round(_e[2]['start'][0],3), round(_e[2]['start'][1],3),
                                    round(_e[2]['end'][0],3),   round(_e[2]['end'][1],3))
                            if _pts not in _seen_pts:
                                _seen_pts.add(_pts)
                                _dedup.append(_e)
                        _edges = _dedup
                        # Tag edges by endpoint connection count (before merge
                        # loses the geometry dict)
                        _edges_tagged = []
                        for _e in _edges:
                            _gln = _e[2]
                            _conn_s, _conn_e = False, False
                            for _pn, _pln in view_parts.items():
                                if _pn in gusset_blk_set:
                                    continue
                                for _ln in _pln:
                                    _d1, _ = dist_pt_to_seg(_gln['start'], _ln['start'], _ln['end'])
                                    _d1 = min(_d1, math.hypot(_gln['start'][0]-_ln['start'][0], _gln['start'][1]-_ln['start'][1]))
                                    _d1 = min(_d1, math.hypot(_gln['start'][0]-_ln['end'][0],   _gln['start'][1]-_ln['end'][1]))
                                    _d2, _ = dist_pt_to_seg(_gln['end'],   _ln['start'], _ln['end'])
                                    _d2 = min(_d2, math.hypot(_gln['end'][0]-_ln['start'][0], _gln['end'][1]-_ln['start'][1]))
                                    _d2 = min(_d2, math.hypot(_gln['end'][0]-_ln['end'][0],   _gln['end'][1]-_ln['end'][1]))
                                    if _d1 <= ADJ_TOL: _conn_s = True
                                    if _d2 <= ADJ_TOL: _conn_e = True
                            _conn = (1 if _conn_s else 0) + (1 if _conn_e else 0)
                            _edges_tagged.append((_e[0], _e[1], _conn))
                        # Merge fragmented colinear edges (polyline-drawn parts)
                        _edges_merged = [(e, op, g_ln) for e, op, g_ln in _edges]  # keep orig for merge
                        if len(_edges_merged) > 1:
                            _n_before = len(_edges_merged)
                            _edges_merged = _merge_collinear_edges(_edges_merged, ADJ_TOL)
                            if len(_edges_merged) < _n_before:
                                print(f"    [merge] reduce gusset edges from {_n_before} to {len(_edges_merged)}")
                        else:
                            _edges_merged = [(e, op, [g_ln]) for e, op, g_ln in _edges_merged]
                        # Map merged edges back to tagged info (sum conn for merged edges)
                        _final = []
                        for _mlen, _mop, _frags in _edges_merged:
                            _all_conn = sum(_c for _l, _op, _c in _edges_tagged if _op == _mop and abs(_l - _mlen) < 1e-6)
                            if _all_conn == 0:
                                _all_conn = max((_c for _l, _op, _c in _edges_tagged if _op == _mop), default=1)
                            _final.append((_mlen, _mop, _all_conn, _frags))
                        # If > expected_edges, drop only truly free edges (conn=0).
                        # conn=1 edges (one endpoint touching) may still be
                        # legitimate weld seams in section views.
                        if len(_final) > expected_edges:
                            _final.sort(key=lambda _x: (_x[2], -_x[0]))
                            while len(_final) > expected_edges and _final[0][2] < 1:
                                _remove = _final.pop(0)
                                print(f"    [free-edge] drop {round(_remove[0]*SCALE,1)}mm (conn={_remove[2]})")
                        _edges = [(_e[0], _e[1], _e[3]) for _e in _final]
                        weld_edges_by_gusset[_gn] = _edges

                # Connected-part enumeration when the 3-SIDES gusset IS the comp body
                # (e.g. BE021).  The gusset's own edges are column construction lines,
                # not weld seams.  Real welds are on the attached non-comp plates.
                if not _synth and part_number_map.get(gusset_name, comp) == comp:
                    _cp_edges = {}
                    for _cpn, _cplns in view_parts.items():
                        if _cpn in gusset_blk_set:
                            continue
                        _cplbl = part_number_map.get(_cpn, comp)
                        if _cplbl == comp:
                            continue
                        # Quick check: does any edge of this part touch the gusset?
                        _touches = False
                        for _cln in _cplns:
                            if _cln['length'] < MIN_EDGE:
                                continue
                            for _gnb in gusset_names:
                                for _gln in view_parts.get(_gnb, []):
                                    _d1, _ = dist_pt_to_seg(_cln['start'], _gln['start'], _gln['end'])
                                    _d2, _ = dist_pt_to_seg(_cln['end'],   _gln['start'], _gln['end'])
                                    if min(_d1, _d2) <= ADJ_TOL:
                                        _touches = True; break
                                if _touches: break
                            if _touches: break
                        if not _touches:
                            continue
                        # Enumerate edges of this connected part
                        _ce_list = []
                        for _cln in _cplns:
                            if _cln['length'] < MIN_EDGE:
                                continue
                            _cp_s = None; _cps_d = ADJ_TOL
                            _cp_e = None; _cpe_d = ADJ_TOL
                            for _opn, _olns in view_parts.items():
                                if _opn == _cpn:
                                    continue
                                for _oln in _olns:
                                    _d1, _ = dist_pt_to_seg(_cln['start'], _oln['start'], _oln['end'])
                                    _d2, _ = dist_pt_to_seg(_cln['end'],   _oln['start'], _oln['end'])
                                    if _d1 <= _cps_d: _cps_d = _d1; _cp_s = _opn
                                    if _d2 <= _cpe_d: _cpe_d = _d2; _cp_e = _opn
                            if not _cp_s and not _cp_e:
                                continue
                            _cln['nb_start'] = _cp_s
                            _cln['nb_end']   = _cp_e
                            _s_gust = _cp_s in gusset_blk_set
                            _e_gust = _cp_e in gusset_blk_set
                            # Use the connected part's block as gusset key,
                            # so output labels use the connected part (e.g. p123).
                            if _cp_s and _cp_e and _cp_s == _cp_e:
                                _ce_list.append((_cln['length'], _cp_s, [_cln]))
                            elif _cp_s and _cp_e:
                                if _s_gust and not _e_gust:
                                    _ce_list.append((_cln['length'], _cp_s, [_cln]))
                                elif _e_gust and not _s_gust:
                                    _ce_list.append((_cln['length'], _cp_e, [_cln]))
                            elif _cp_s and _s_gust:
                                _ce_list.append((_cln['length'], _cp_s, [_cln]))
                            elif _cp_e and _e_gust:
                                _ce_list.append((_cln['length'], _cp_e, [_cln]))
                        if _ce_list:
                            _cp_edges[_cpn] = _ce_list
                    if _cp_edges:
                        weld_edges_by_gusset = _cp_edges

                weld_edges_all = [(e, op, gn, frags)
                                  for gn, edges in weld_edges_by_gusset.items()
                                  for e, op, frags in edges]
                if not weld_edges_all:
                    if is_circle_wm:
                        skipped.append((wm_name, "CIRCLE: no adjacent edges found"))
                    else:
                        skipped.append((wm_name, "3 SIDES: no adjacent edges found"))
                    continue

                lbl_g = part_number_map.get(gusset_name, comp)
                _tag3 = "CIRCLE" if is_circle_wm else "3 SIDES"
                _wm_short = wm_name.split(' - ')[0]
                print(f"  [{view_id}] {_wm_short}  [{_tag3}]  gusset={lbl_g}")
                # TYP multiplier: count candidate non-comp labels in the main assembly view.
                # Uses the gusset label and all other-part labels from collected edges.
                typ_mul_3s = 1
                if parsed['is_typ']:
                    _cand = {part_number_map.get(op, comp)
                             for _, op, _, _ in weld_edges_all
                             if part_number_map.get(op, comp) != comp}
                    _cand.add(lbl_g)
                    _cand.discard(comp)
                    if _cand:
                        typ_mul_3s = max(
                            sum(1 for k, v in part_number_map.items()
                                if v == cl and k.split(' - ')[-1] == main_view_id)
                            for cl in _cand
                        )
                    # Divide by gusset count: multi-gusset logic already covers
                    # repetition when len(gusset_names) > 1, so avoid double-counting.
                    typ_mul_3s = max(1, typ_mul_3s // len(gusset_names))
                    if typ_mul_3s > 1:
                        print(f"    [TYP x{typ_mul_3s}]")
                # hf correction for 3-SIDES: skip when CJP annotation is present
                # or both sides have the same valid fillet size.
                _sz3_a = parsed['size_above']
                _sz3_b = parsed['size_below']
                if (parsed['groove_above'] or parsed['groove_below']
                        or (_sz3_a is not None and _sz3_a == _sz3_b and _sz3_a <= MAX_HF)):
                    sz3_above = _sz3_a
                    sz3_below = _sz3_b
                else:
                    sz3_above = _correct_hf_3s(parsed['size_above'], lbl_g)
                    sz3_below = _correct_hf_3s(parsed['size_below'], lbl_g)

                # Rank-based BOM mapping for 3-SIDES gussets with known dimensions.
                # Two strategies depending on edge-length distribution:
                #
                #   A) 2 distinct lengths with one appearing twice (e.g. p42 edges
                #      [33, 120, 120]): the duplicated length maps to whichever BOM
                #      dim is closer; the singleton maps to the other BOM dim.
                #
                #   B) 3 distinct lengths: sort ascending and pair with sorted
                #      [smaller_BOM_dim, smaller_BOM_dim, larger_BOM_dim].
                #
                # Applied only when gusset is a non-comp plate and at least one
                # geo edge is within 25 % of one BOM dimension.
                _bom_edge_map = {}
                if lbl_g != comp and lbl_g in part_dims:
                    _pd3 = part_dims[lbl_g]
                    _bw3 = _pd3.get('width')
                    _bl3 = _pd3.get('bom_len')
                    # If BOM has width but no length, estimate from the
                    # gusset's median geo edge (e.g. BE022 p200 268mm).
                    if _bw3 and not _bl3:
                        _geo_vals = sorted(set(
                            round(_el * SCALE, 1)
                            for _el, _op, _cg, _ in weld_edges_all
                            if part_number_map.get(_cg, comp) == lbl_g
                        ))
                        if len(_geo_vals) >= 2:
                            _bl3 = _geo_vals[len(_geo_vals)//2]  # median
                            print(f"    [BOM infer] {lbl_g} L={_bl3} (from geo median {_geo_vals})")
                    
                    # Strategy C: Flange plate override (for p200-like plates)
                    # Check if this is a flange plate (width ≈ comp flange width)
                    _is_flange_plate = False
                    if _bw3:
                        # Check against comp flange width if available
                        if comp_dims.get('flange_w') and abs(_bw3 - comp_dims['flange_w']) < 10:
                            _is_flange_plate = True
                        # Also check for typical flange plate widths (140mm for H300, etc.)
                        elif _bw3 in [140, 145, 150]:
                            _is_flange_plate = True
                    if _bw3 and _is_flange_plate:
                        # Collect all unique geo lengths from this gusset
                        _gusset_geo_lens = []
                        for _el, _op, _cg, _ in weld_edges_all:
                            _geo_mm = round(_el * SCALE, 1)
                            if _geo_mm not in _gusset_geo_lens:
                                _gusset_geo_lens.append(_geo_mm)
                        
                        # Check if geo edges are far from BOM width (section-view distortion)
                        _all_far = all(
                            abs(_g - _bw3) / max(_g, 1) > 0.4
                            for _g in _gusset_geo_lens
                        )
                        
                        if _all_far and len(_gusset_geo_lens) >= 2:
                            # Map: largest geo → comp depth, others → plate width
                            _sorted_geo = sorted(_gusset_geo_lens)
                            _comp_depth = comp_dims.get('depth', _bl3 if _bl3 else 270)
                            
                            # Map all geo edges for all gusset instances
                            for _cg in set(_cg for _, _, _cg, _ in weld_edges_all):
                                for _g in _gusset_geo_lens:
                                    if _g == _sorted_geo[-1]:
                                        # Largest edge → comp depth
                                        _bom_edge_map[(_cg, _g)] = round(_comp_depth)
                                    else:
                                        # Other edges → plate width
                                        _bom_edge_map[(_cg, _g)] = round(_bw3)
                            
                            print(f"    [BOM map-flange] {lbl_g}  w={_bw3} depth={round(_comp_depth)} (geo far from BOM)")
                    
                    if _bw3 and _bl3:
                        _bom_dims = sorted([_bw3, _bl3])  # [smaller, larger]
                        # Collect dedup edge lengths per gusset WITH multiplicity
                        _gusset_geo_counts = defaultdict(Counter)
                        _seenv = set()
                        for _el, _op, _cg, _ in weld_edges_all:
                            _bp = (_cg, _op, round(_el, 2))
                            if _bp not in _seenv:
                                _seenv.add(_bp)
                                _gusset_geo_counts[_cg][round(_el * SCALE, 1)] += 1
                        for _cg, _geo_counter in _gusset_geo_counts.items():
                            _total = sum(_geo_counter.values())
                            if _total < 2:
                                continue
                            _any_match = any(
                                min(abs(_g - d) / max(_g, 1) for d in _bom_dims) < 0.25
                                for _g in _geo_counter
                            )
                            if not _any_match:
                                continue
                            if _total == 3 and len(_geo_counter) == 2:
                                # Strategy A — duplicate length + singleton
                                _dup_len = max(_geo_counter, key=_geo_counter.get)
                                _uniq_len = min(_geo_counter, key=_geo_counter.get)
                                _d0 = abs(_dup_len - _bom_dims[0])
                                _d1 = abs(_dup_len - _bom_dims[1])
                                # Skip when distances to both BOM dims are too close
                                # (ambiguous assignment — e.g. p126 geo=90.5 vs bw=110 bl=116,
                                #  Δ=19.5 vs 25.5, diff=6.6% → keep geo)
                                if abs(_d0 - _d1) / max(_dup_len, 1) < 0.08:
                                    pass  # ambiguous, keep geo
                                elif _d0 <= _d1:
                                    _bom_edge_map[(_cg, _dup_len)] = round(_bom_dims[0])
                                    # Only map singleton if reasonably close
                                    if abs(_uniq_len - _bom_dims[1]) / max(_uniq_len, 1) < 0.40:
                                        _bom_edge_map[(_cg, _uniq_len)] = round(_bom_dims[1])
                                else:
                                    _bom_edge_map[(_cg, _dup_len)] = round(_bom_dims[1])
                                    if abs(_uniq_len - _bom_dims[0]) / max(_uniq_len, 1) < 0.40:
                                        _bom_edge_map[(_cg, _uniq_len)] = round(_bom_dims[0])
                            elif _total == 3 and len(_geo_counter) == 3:
                                # Strategy B — three unique lengths
                                # BOM pattern is [bw, bl] → 3 edges = [bw, bw, bl]
                                # Match the edge closest to bl (longer BOM dim) to length,
                                # and the other two to width.  Avoids the positional
                                # sort-and-pair pitfall (e.g. geo 231/269/439 with
                                # bw=140 bl=268 should map 269→268 not 269→140).
                                _geo_sorted = sorted(_geo_counter.elements())
                                _bw_smaller = _bom_dims[0]
                                _bl_larger  = _bom_dims[1]
                                _dists_to_bl = [
                                    abs(g - _bl_larger) / max(g, 1)
                                    for g in _geo_sorted
                                ]
                                _best_bl_idx = _dists_to_bl.index(min(_dists_to_bl))
                                for _i, _geo in enumerate(_geo_sorted):
                                    if _i == _best_bl_idx:
                                        _bom_edge_map[(_cg, _geo)] = round(_bl_larger)
                                    else:
                                        # Only map to bw if the edge is within 80%
                                        # range (avoids e.g. p42 33→73 but allows
                                        # p200 438.9→140 in section-view projection)
                                        if abs(_geo - _bw_smaller) / max(_geo, 1) < 0.80:
                                            _bom_edge_map[(_cg, _geo)] = round(_bw_smaller)
                        if _bom_edge_map:
                            print(f"    [BOM map] {lbl_g}  w={_bw3} L={_bl3}")

                # Dedup by (gusset_block, other_block, edge_len): prevents counting the
                # same physical line twice (symmetric DXF where one Part appears in two
                # views at identical positions) while allowing distinct gusset instances
                # (e.g. two haunch blocks at different positions) to each contribute edges.
                edge_rows = []   # accumulate edge rows; extended by typ_mul_3s at end
                seen_bp = set()
                for edge_len, other_part, cur_gusset, edge_frags in weld_edges_all:
                    _bp = (cur_gusset, other_part, round(edge_len, 2))
                    if not _use_largest_gusset and _bp in seen_bp:
                        continue
                    seen_bp.add(_bp)
                    lbl_o       = part_number_map.get(other_part, comp)
                    geo_len_mm  = round(edge_len * SCALE, 1)
                    _lbl_g_dedup = part_number_map.get(cur_gusset, comp)
                    # Fragment-level label override
                    if lbl_o == comp:
                        _fnc = {}
                        _fcomp = False
                        for g_ln in edge_frags:
                            _ns = g_ln.get('nb_start')
                            _ne = g_ln.get('nb_end')
                            if _ns and _ne and _ns != _ne:
                                continue
                            for _nb in (_ns, _ne):
                                if _nb and _nb not in gusset_blk_set:
                                    _nbl = part_number_map.get(_nb, comp)
                                    if _nbl == comp:
                                        _fcomp = True
                                    elif _nbl != _lbl_g_dedup:
                                        _fnc[_nbl] = _fnc.get(_nbl,0)+1
                        if _fnc and not _fcomp:
                            lbl_o = max(_fnc, key=_fnc.get)
                            print(f'    [frag ovr] {other_part}->{lbl_o} (nb data)')
                    # Per-edge label override: when other_part is unlabeled (maps to comp),
                    # scan source fragments for a closer non-comp neighbour.
                    # Skip for synthetic edges (dummy fragments have no real geometry).
                    _lbl_g_dedup = part_number_map.get(cur_gusset, comp)
                    if not _use_largest_gusset and lbl_o == comp:
                        _comp_d = 1e9
                        _best_nc = None
                        _best_nc_d = 1e9
                        for g_ln in edge_frags:
                            for ep in (g_ln['start'], g_ln['end']):
                                for pn, plns in view_parts.items():
                                    if pn in gusset_blk_set:
                                        continue
                                    _pn_lbl = part_number_map.get(pn, comp)
                                    for ln in plns:
                                        d = min(
                                            math.hypot(ep[0]-ln['start'][0], ep[1]-ln['start'][1]),
                                            math.hypot(ep[0]-ln['end'][0], ep[1]-ln['end'][1]),
                                        )
                                        d_int, _ = dist_pt_to_seg(ep, ln['start'], ln['end'])
                                        d = min(d, d_int)
                                        if _pn_lbl == comp:
                                            if d < _comp_d:
                                                _comp_d = d
                                        elif _pn_lbl != _lbl_g_dedup:
                                            if d < _best_nc_d:
                                                _best_nc_d = d
                                                _best_nc = _pn_lbl
                        if _best_nc and (_best_nc_d <= _comp_d or _comp_d > ADJ_TOL):
                            if _best_nc == comp or _lbl_g_dedup == comp:
                                lbl_o = _best_nc
                                print(f"    [per-edge ovr] {other_part}->{_best_nc} nc_d={round(_best_nc_d,1)} comp_d={round(_comp_d,1)}")
                    # Cross-view dedup: same gusset label + same other label + same
                    # geo length in a DIFFERENT view → same physical edge shown twice.
                    # Only when BOM qty == 1 (single-instance; multi-instance parts
                    # in different views are different physical copies).
                    _bom_qty_g = part_dims.get(_lbl_g_dedup, {}).get('qty', 1)
                    _cur_vid = cur_gusset.split(' - ')[-1] if ' - ' in cur_gusset else view_id
                    _lbl_key = (_lbl_g_dedup, lbl_o, geo_len_mm)
                    if (_bom_qty_g == 1 and _lbl_key in cross_view_seen
                            and cross_view_seen[_lbl_key] != _cur_vid):
                        continue
                    cross_view_seen[_lbl_key] = _cur_vid
                    # Priority 1: rank-based BOM mapping (3 edges → [W, W, L])
                    final_edge_mm = _bom_edge_map.get((cur_gusset, geo_len_mm), None)
                    if final_edge_mm is not None:
                        pass  # BOM rank mapping applied
                    else:
                        final_edge_mm = geo_len_mm
                        # Priority 2: single-edge bom_len correction
                        if lbl_g in part_dims:
                            _pd3 = part_dims[lbl_g]
                            _bw3 = _pd3.get('width')
                            _bl3 = _pd3.get('bom_len')
                            if (_bw3 and _bl3 and geo_len_mm > 0
                                    and abs(geo_len_mm - _bl3) / geo_len_mm < 0.15
                                    and abs(geo_len_mm - _bw3) / geo_len_mm > 0.35):
                                final_edge_mm = round(_bl3)
                    # Normalize: comp in part1; if neither is comp, gusset in part1
                    if lbl_o == comp:
                        p1, p2 = lbl_o, lbl_g
                    elif lbl_g == comp:
                        p1, p2 = lbl_g, lbl_o
                    else:
                        p1, p2 = (lbl_g, lbl_o) if lbl_g <= lbl_o else (lbl_o, lbl_g)
                    print(f"    edge geo={geo_len_mm}mm final={final_edge_mm}mm  {p1}/{p2}")
                    if p1 == p2 and not (comp.startswith('CO') and p1 == comp):
                        # self-reference; CO components legitimately have
                        # CO/CO pairs (stiffener->column body) but non-comp
                        # self-references (e.g. p144/p144) are still skipped.
                        continue
                    # CJP normalization for 3-SIDES edges (same rule as normal WMs)
                    grove_3s_ab = parsed['groove_above']
                    grove_3s_bl = parsed['groove_below']
                    s3_data = []
                    for side, sz, present, is_g in [
                        ('Above', sz3_above, parsed['has_above'] or grove_3s_ab, grove_3s_ab),
                        ('Below', sz3_below, parsed['has_below'] or grove_3s_bl, grove_3s_bl),
                    ]:
                        if present:
                            s3_data.append({'side': side, 'sz': sz, 'is_groove': is_g})
                    cjp3 = [s for s in s3_data if s['is_groove']]
                    fil3 = [s for s in s3_data if not s['is_groove']]
                    if cjp3:
                        edge_rows.append({
                            'component': comp, 'position': 'Above',
                            'hf': None, 'length_mm': final_edge_mm,
                            'annotation': 'CJP', 'part1': p1, 'part2': p2,
                        })
                        if fil3:
                            f3 = fil3[0]
                            _hf_fb = 0
                            if f3['sz'] is not None:
                                _hf_fb = f3['sz']
                            elif lbl_g in part_dims:
                                _hf_fb = hf_from_thickness(part_dims[lbl_g]['thick'])
                            elif comp_web_t:
                                _hf_fb = hf_from_thickness(comp_web_t)
                            else:
                                _hf_fb = 6  # default minimum fillet
                            edge_rows.append({
                                'component': comp, 'position': 'Below',
                                'hf': _hf_fb,
                                'length_mm': final_edge_mm,
                                'annotation': '', 'part1': p1, 'part2': p2,
                            })
                    else:
                        for s3 in s3_data:
                            _hf_fb = 0
                            if s3['sz'] is not None:
                                _hf_fb = s3['sz']
                            elif lbl_g in part_dims:
                                _hf_fb = hf_from_thickness(part_dims[lbl_g]['thick'])
                            elif comp_web_t:
                                _hf_fb = hf_from_thickness(comp_web_t)
                            else:
                                _hf_fb = 6
                            edge_rows.append({
                                'component': comp, 'position': s3['side'],
                                'hf': _hf_fb, 'length_mm': final_edge_mm,
                                'annotation': '', 'part1': p1, 'part2': p2,
                            })
                results.extend(edge_rows * typ_mul_3s)
                continue  # skip normal weld processing for 3-SIDES

            # ---- Normal weld ----
            best_part, weld_line, match_how = choose_weld_line(arrow, matches)
            weld_len_mm = round(weld_line['length'] * SCALE, 1)

            other_parts = [m['part'] for m in matches if m['part'] != best_part]
            part2_name  = other_parts[0] if other_parts else None

            lbl1 = part_number_map.get(best_part, comp)
            lbl2 = part_number_map.get(part2_name, comp) if part2_name else comp
            if lbl1 == lbl2:
                lbl2 = comp
            # Normalize order: comp always in part1; other pairs sorted alphabetically
            if lbl2 == comp and lbl1 != comp:
                lbl1, lbl2 = lbl2, lbl1
            elif lbl1 != comp and lbl2 != comp and lbl1 > lbl2:
                lbl1, lbl2 = lbl2, lbl1
            # Comp-backoff: when both labels are non-comp and the comp has
            # no labelled Part block in this view (common in column section
            # cuts), replace the nearest match with comp.
            if lbl1 != comp and lbl2 != comp:
                _has_comp = any(
                    v == comp
                    for k, v in part_number_map.items()
                    if k.split(' - ')[-1] == view_id
                )
                if not _has_comp:
                    lbl1 = comp

            lbl_non_comp = lbl2 if lbl1 == comp else lbl1
            bom_fallback_count = 1

            # BOM fallback: when the WM finds only comp-labeled parts (self-weld),
            # the non-comp plate is not visible in the elevation view.  Scan BOM
            # for a part whose bom_width ≈ geo (within 15 %) to recover the label.
            if lbl_non_comp == comp and part_dims and weld_len_mm > 0:
                _best_ratio = 0.15
                _best_lbl   = None
                for _plbl, _pdims in part_dims.items():
                    if _plbl == comp:
                        continue
                    _bw = _pdims.get('width')
                    if _bw and _bw > 0:
                        _r = abs(weld_len_mm - _bw) / weld_len_mm
                        if _r < _best_ratio:
                            _best_ratio = _r
                            _best_lbl   = _plbl
                if _best_lbl:
                    lbl2 = _best_lbl
                    lbl_non_comp = _best_lbl
                    bom_fallback_count = sum(
                        1 for lbl in part_number_map.values()
                        if lbl == _best_lbl
                    )

            # TYP multiplier: use BOM qty for column-type components where
            # stiffeners appear in separate section views (not all in main).
            if parsed['is_typ'] and lbl_non_comp != comp:
                _bom_qty = part_dims.get(lbl_non_comp, {}).get('qty', 1)
                _view_n  = sum(1 for k, v in part_number_map.items()
                               if v == lbl_non_comp and k.split(' - ')[-1] == main_view_id)
                if comp.startswith('CO') and _bom_qty and _bom_qty > _view_n:
                    _typ_n = _bom_qty  # BOM is more reliable for column stiffeners
                else:
                    _typ_n = _view_n
                if _typ_n > 1:
                    bom_fallback_count = _typ_n
                    print(f"    [TYP x{bom_fallback_count}] {lbl_non_comp}")

            # Stiffener flange-face override (any match type):
            # When the non-comp plate width ≈ comp flange width (cover/stiffener plate
            # spanning the full flange) AND hf ≥ 10 mm (flange-face weld), the weld
            # length equals the plate width — regardless of which line was geometrically
            # selected (which may be the comp's flange/web line instead of the plate).
            sz_above_raw = parsed['size_above']
            sz_below_raw = parsed['size_below']
            max_hf_raw = max(
                sz_above_raw if sz_above_raw is not None else 0,
                sz_below_raw if sz_below_raw is not None else 0,
            )
            stiffener_override_applied = False
            if (lbl_non_comp != comp
                    and lbl_non_comp in part_dims
                    and comp_dims.get('flange_w')
                    and max_hf_raw >= 10):
                pd = part_dims[lbl_non_comp]
                if abs(pd['width'] - comp_dims['flange_w']) < 5:
                    weld_len_mm = round(pd['width'])
                    stiffener_override_applied = True

            # BOM-width correction (any match type):
            # Three cases where BOM dimensions override the geometry length:
            #   Case 1: geo ≈ bom_len → plate end-face weld → use bom_width
            #           (plate drawn along its length, weld on end face).
            #   Case 2: geo ≈ bom_width, bom_len far from bom_width
            #           → plate width is short dimension, weld runs along bom_len.
            #   Case 3: geo ≈ bom_width within 25 % → section-view approximation.
            # Skipped when the stiffener override already set the length.
            # All BOM widths are rounded to the nearest mm (engineering convention).
            BOM_WIDTH_TOL = 0.25
            BOM_LEN_TOL   = 0.08
            weld_len_mm_orig = weld_len_mm  # save for CO fallback logging
            print(f"    [BOM pre-check] lbl_nc={lbl_non_comp} stiff={stiffener_override_applied} in_dims={lbl_non_comp in part_dims} wlm={weld_len_mm}")
            if (not stiffener_override_applied
                    and lbl_non_comp != comp
                    and lbl_non_comp in part_dims):
                pd_nc = part_dims[lbl_non_comp]
                bw = pd_nc['width']
                bl = pd_nc.get('bom_len')
                if bw and bw > 0 and weld_len_mm > 0:
                    if bl and bl > 0 and abs(weld_len_mm - bl) / weld_len_mm < BOM_LEN_TOL:
                        # Case 1: geo ≈ bom_len
                        # Sub-case: if geo also matches bw closely, prefer bl (both dimensions match)
                        if abs(weld_len_mm - bw) / max(weld_len_mm, 1) < BOM_LEN_TOL:
                            # Both bw and bl match; prefer bl (typically the weld
                            # run length, as engineering convention rounds up)
                            print(f"    [BOM case1-both] {lbl_non_comp} geo={weld_len_mm} bw={bw} bl={bl}")
                            weld_len_mm = round(bl)
                        elif abs(bl - bw) / max(bl, 1) > 0.3:
                            # bl and bw are very different (not a square plate)
                            # geo matches bl → weld runs along plate length, keep geo unchanged
                            # This handles cases like p26: geo=200, bw=95, bl=200
                            print(f"    [BOM case1-skip] {lbl_non_comp} geo={weld_len_mm} bw={bw} bl={bl} (geo=bl, keep geo)")
                        elif abs(weld_len_mm - bw) / max(weld_len_mm, 1) > 0.3:
                            # Only bl matches and geo far from bw → plate end-face weld, use bw
                            print(f"    [BOM case1] {lbl_non_comp} geo={weld_len_mm} bw={bw} bl={bl}")
                            weld_len_mm = round(bw)
                        else:
                            # geo ≈ bl but also somewhat close to bw → keep geo (weld along length)
                            print(f"    [BOM case1-skip2] {lbl_non_comp} geo={weld_len_mm} bw={bw} bl={bl} (geo close to both, keep geo)")
                    elif (bl and bl > 0
                          and abs(weld_len_mm - bw) / weld_len_mm < 0.05
                          and abs(bl - bw) / max(bw, 1) > 0.3):
                        # Case 2: geo ≈ bom_width closely, but bom_len is a
                        # different dimension → weld runs along bom_len
                        print(f"    [BOM case2] {lbl_non_comp} geo={weld_len_mm} bw={bw} bl={bl}")
                        weld_len_mm = round(bl)
                    elif abs(weld_len_mm - bw) / weld_len_mm < BOM_WIDTH_TOL:
                        # Case 3: geo close to bom_width
                        print(f"    [BOM case3] {lbl_non_comp} geo={weld_len_mm} bw={bw}")
                        weld_len_mm = round(bw)
                    else:
                        print(f"    [BOM no-case] {lbl_non_comp} geo={weld_len_mm} bw={bw} bl={bl}")
            # CO section-view fallback: column-type section cuts show plates
            # in foreshortened projection (e.g. p124 geo=90.5 → bw=116,
            # geo=170 → bl=220).  Standard beam tolerances are too strict.
            if (comp.startswith('CO')
                    and not stiffener_override_applied
                    and lbl_non_comp != comp
                    and lbl_non_comp in part_dims
                    and weld_len_mm > 0):
                pd_nc = part_dims[lbl_non_comp]
                _bw = pd_nc['width']
                _bl = pd_nc.get('bom_len')
                if _bw and _bw > 0 and _bl and _bl > 0:
                    _dw = abs(weld_len_mm - _bw) / weld_len_mm
                    _dl = abs(weld_len_mm - _bl) / weld_len_mm
                    if _dw < 0.35 and _dw < _dl:
                        weld_len_mm = round(_bw)
                        print(f"    [BOM co-fallback] {lbl_non_comp} geo={weld_len_mm_orig}→bw={weld_len_mm}")
                    elif _dl < 0.35 and _dl < _dw:
                        weld_len_mm = round(_bl)
                        print(f"    [BOM co-fallback] {lbl_non_comp} geo={weld_len_mm_orig}→bl={weld_len_mm}")

            final_len_mm = weld_len_mm

            # hf correction: skip when CJP/groove is present OR when both
            # sides have the same valid fillet size (double-sided fillet —
            # the size is clearly a deliberate weld annotation, not a
            # plate-thickness proxy accidentally matching the web/flange).
            _sz_a = parsed['size_above']
            _sz_b = parsed['size_below']
            if (parsed['groove_above'] or parsed['groove_below']
                    or (_sz_a is not None and _sz_a == _sz_b and _sz_a <= MAX_HF)):
                sz_above = _sz_a
                sz_below = _sz_b
            else:
                sz_above = _correct_hf(parsed['size_above'], lbl1, lbl2)
                sz_below = _correct_hf(parsed['size_below'], lbl1, lbl2)

            print(f"  [{view_id}] {wm_name.split(' - ')[0]}")
            print(f"    arrow={arrow}  geo={weld_len_mm}mm  final={final_len_mm}mm"
                  f"  parts: {lbl1} / {lbl2}"
                  f"  size\u2191{sz_above} \u2193{sz_below}"
                  f"  annot={parsed['annotation']!r}")

            if lbl1 == lbl2 and not (comp.startswith('CO') and lbl1 == comp):
                # self-reference; CO components legitimately have CO/CO
                # pairs (stiffener->column body).
                continue

            # CJP/groove normalization:
            # CJP always output as position='Above' with hf=None and note='CJP'.
            # The paired fillet (if any) is output as position='Below' with hf=value.
            # For pure CJP (no paired fillet) only one row is emitted.
            groove_ab = parsed['groove_above']
            groove_bl = parsed['groove_below']
            sides_data = []
            for side, sz, present, is_groove in [
                ('Above', sz_above, parsed['has_above'] or groove_ab, groove_ab),
                ('Below', sz_below, parsed['has_below'] or groove_bl, groove_bl),
            ]:
                if present:
                    sides_data.append({'side': side, 'sz': sz, 'is_groove': is_groove})

            cjp_sides    = [s for s in sides_data if s['is_groove']]
            fillet_sides = [s for s in sides_data if not s['is_groove']]

            if cjp_sides:
                # CJP side → always 'Above', hf=None, note='CJP'
                for _rep in range(bom_fallback_count):
                    results.append({
                        'component':  comp,
                        'position':   'Above',
                        'hf':         None,
                        'length_mm':  final_len_mm,
                        'annotation': 'CJP',
                        'part1':      lbl1,
                        'part2':      lbl2,
                    })
                if fillet_sides:
                    # Paired fillet → 'Below', hf=value (fallback to standard if needed)
                    f = fillet_sides[0]
                    _hf_fb = f['sz'] if f['sz'] is not None else 0
                    if _hf_fb == 0:
                        if lbl_non_comp in part_dims:
                            _hf_fb = hf_from_thickness(part_dims[lbl_non_comp]['thick'])
                        elif comp_web_t:
                            _hf_fb = hf_from_thickness(comp_web_t)
                        else:
                            _hf_fb = 6
                    for _rep in range(bom_fallback_count):
                        results.append({
                            'component':  comp,
                            'position':   'Below',
                            'hf':         _hf_fb,
                            'length_mm':  final_len_mm,
                            'annotation': '',
                            'part1':      lbl1,
                            'part2':      lbl2,
                        })
            else:
                # Normal fillet: output each side as-is
                for side, sz, present, _ in [
                    ('Above', sz_above, parsed['has_above'], None),
                    ('Below', sz_below, parsed['has_below'], None),
                ]:
                    if not present:
                        continue
                    hf_val = sz if sz is not None else 0
                    if hf_val == 0:
                        if lbl_non_comp in part_dims:
                            hf_val = hf_from_thickness(part_dims[lbl_non_comp]['thick'])
                        elif comp_web_t:
                            hf_val = hf_from_thickness(comp_web_t)
                        else:
                            hf_val = 6
                    for _rep in range(bom_fallback_count):
                        results.append({
                            'component':  comp,
                            'position':   side,
                            'hf':         hf_val,
                            'length_mm':  final_len_mm,
                            'annotation': '',
                            'part1':      lbl1,
                            'part2':      lbl2,
                    })

    # Post-processing: connected-part enumeration for 3-SIDES views
    # where gusset is the comp body. Only for BE (non-CO) components.
    if not comp.startswith('CO') and part_lines_map and part_dims:
        _ADJ = SNAP_TOL + 0.5
        _MIN_EDGE_CAD = 1.5
        _plates_done = set()
        for r in results:
            _plates_done.add(r['part1'])
            _plates_done.add(r['part2'])
        for _vid, _vparts in part_lines_map.items():
            # Only run for views with 3-SIDES WeldMarks
            if not any(any(kw in (parse_weldmark(_wb) or {}).get('annotation','').upper()
                          for kw in ('SIDE','\u56f4','\u5168'))
                      for _wn, _wb in wm_by_view.get(_vid, [])):
                continue
            _comp_blocks = {pn for pn in _vparts if part_number_map.get(pn, comp) == comp}
            if not _comp_blocks:
                continue
            for _cpn, _cplns in _vparts.items():
                _cplbl = part_number_map.get(_cpn, comp)
                if _cplbl == comp or _cplbl in _plates_done:
                    continue
                # Check if this part touches the comp body
                _touches = False
                for _cln in _cplns:
                    if _cln['length'] <= _MIN_EDGE_CAD:
                        continue
                    for _cbn in _comp_blocks:
                        for _gln in _vparts.get(_cbn, []):
                            _d1, _ = dist_pt_to_seg(_cln['start'], _gln['start'], _gln['end'])
                            _d2, _ = dist_pt_to_seg(_cln['end'],   _gln['start'], _gln['end'])
                            if min(_d1, _d2) <= _ADJ:
                                _touches = True; break
                        if _touches: break
                    if _touches: break
                if not _touches:
                    continue
                _t = part_dims.get(_cplbl, {}).get('thick', 12)
                _hf = hf_from_thickness(_t) if _t else 8
                for _cln in _cplns:
                    if _cln['length'] <= _MIN_EDGE_CAD:
                        continue
                    _cp_s = None; _cps_d = _ADJ
                    _cp_e = None; _cpe_d = _ADJ
                    for _opn, _olns in _vparts.items():
                        if _opn == _cpn:
                            continue
                        for _oln in _olns:
                            _d1, _ = dist_pt_to_seg(_cln['start'], _oln['start'], _oln['end'])
                            _d2, _ = dist_pt_to_seg(_cln['end'],   _oln['start'], _oln['end'])
                            if _d1 <= _cps_d: _cps_d = _d1; _cp_s = _opn
                            if _d2 <= _cpe_d: _cpe_d = _d2; _cp_e = _opn
                    if not _cp_s and not _cp_e:
                        continue
                    _nbl_s = part_number_map.get(_cp_s, comp)
                    _nbl_e = part_number_map.get(_cp_e, comp) if _cp_e else comp
                    if _cp_s in _comp_blocks and _cp_e in _comp_blocks:
                        continue  # both endpoints touch comp → design envelope, not weld
                    elif _cp_s in _comp_blocks:
                        _lbl_o = comp  # prefer comp over non-comp
                    elif _cp_e in _comp_blocks:
                        _lbl_o = comp
                    elif _cp_s and _cp_e and _cp_s == _cp_e:
                        _lbl_o = _nbl_s  # both touch same non-comp → plate→plate
                    _wlen = round(_cln['length'] * SCALE, 1)
                    _pair = tuple(sorted((_cplbl, _lbl_o)))
                    if _pair == tuple(sorted((comp, _cplbl))) or _pair == tuple(sorted((_cplbl, comp))):
                        p1, p2 = comp, _cplbl
                    else:
                        p1, p2 = _pair
                    if p1 == p2:
                        continue
                    for _pos in ('Above', 'Below'):
                        for _dup in (0, 1):  # ×2 for symmetric plate faces
                            results.append({
                                'component': comp, 'position': _pos,
                                'hf': _hf, 'length_mm': _wlen,
                                'annotation': '', 'part1': p1, 'part2': p2,
                            })

    if skipped:
        print(f"\n  SKIPPED ({len(skipped)}):")
        for name, reason in skipped:
            print(f"    {name}: {reason}")

    print(f"  → {len(results)} weld rows")
    return results, skipped

# ============================================================
# Excel output
# ============================================================
def write_excel(all_results, all_skipped, output_path):
    wb = openpyxl.Workbook()

    # ---- Sheet 1: Weld statistics ----
    ws = wb.active
    ws.title = "焊缝统计"

    HDR_FILL = PatternFill("solid", fgColor="4472C4")
    HDR_FONT = Font(bold=True, color="FFFFFF")
    CENTER    = Alignment(horizontal='center', vertical='center')

    headers = ['序号', '位置(上/下)', '焊脚尺寸hf(mm)', '焊缝长度(mm)',
               '备注', '零件1', '零件2', '构件号']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = HDR_FILL
        cell.font = HDR_FONT
        cell.alignment = CENTER

    for idx, r in enumerate(all_results, 1):
        ws.cell(row=idx+1, column=1, value=idx)
        ws.cell(row=idx+1, column=2, value=r['position'])
        ws.cell(row=idx+1, column=3, value=r['hf'])
        ws.cell(row=idx+1, column=4, value=r['length_mm'])
        ws.cell(row=idx+1, column=5, value=r['annotation'])
        ws.cell(row=idx+1, column=6, value=r['part1'])
        ws.cell(row=idx+1, column=7, value=r['part2'])
        ws.cell(row=idx+1, column=8, value=r['component'])

    for col in ws.columns:
        w = max((len(str(cell.value or '')) for cell in col), default=0)
        ws.column_dimensions[col[0].column_letter].width = max(w + 3, 14)

    # ---- Sheet 2: Skipped / errors ----
    ws2 = wb.create_sheet("异常报告")
    ws2.cell(row=1, column=1, value="WeldMark 名称").font = Font(bold=True)
    ws2.cell(row=1, column=2, value="原因").font = Font(bold=True)
    for idx, (name, reason) in enumerate(all_skipped, 2):
        ws2.cell(row=idx, column=1, value=name)
        ws2.cell(row=idx, column=2, value=reason)
    ws2.column_dimensions['A'].width = 50
    ws2.column_dimensions['B'].width = 40

    wb.save(output_path)
    print(f"\nSaved → {output_path}")
    print(f"Total weld rows : {len(all_results)}")
    print(f"Total skipped   : {len(all_skipped)}")

# ============================================================
# Entry point
# ============================================================
if __name__ == '__main__':
    dxf_files = sorted(glob.glob(os.path.join(FOLDER, "*.dxf")))
    if not dxf_files:
        print("No DXF files found. Run convert_dwg_to_dxf.py first.")
        raise SystemExit(1)

    all_results = []
    all_skipped = []

    for dxf_path in dxf_files:
        try:
            results, skipped = extract_welds(dxf_path)
            all_results.extend(results)
            all_skipped.extend(skipped)
        except Exception as exc:
            import traceback
            print(f"\nERROR: {dxf_path}\n{traceback.format_exc()}")

    write_excel(all_results, all_skipped, OUTPUT)
